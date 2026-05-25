from __future__ import annotations

from dataclasses import dataclass
from math import sin
from pathlib import Path
from random import Random

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
ONEDRIVE_FILE = Path(
    "/Users/alejandroalvear/Library/CloudStorage/OneDrive-Finning/"
    "SAP_Estudio/resiliencia_financiera_tesis_mafi.xlsx"
)
LOCAL_FILE = BASE_DIR / "data" / "processed" / "resiliencia_financiera_tesis_mafi.xlsx"


@dataclass(frozen=True)
class Client:
    id_cliente: str
    codigo_cliente_sap: str
    nombre_cliente: str
    industria: str
    grupo_economico: str
    segmento: str
    faena_operacion: str
    region: str


@dataclass(frozen=True)
class Product:
    id_producto_equipo: str
    codigo_material_sap: str
    familia_producto: str
    subfamilia_producto: str
    linea_negocio: str
    tipo_equipo: str
    modelo: str
    comercializacion_equipo: str
    estado_ciclo_vida: str
    vida_util_esperada_meses: int


def main() -> None:
    rng = Random(42)
    periods = _periods()
    clients = _clients()
    products = _products()
    cells = _cells(clients, products, rng)
    gav = _gav_monthly(periods)
    facts = _facts(periods, cells, clients, products, gav, rng)
    relationship_cells = _relationship_cells(cells, facts)
    matrix = _matrix(cells, clients, products, facts)
    family_summary = _family_summary(facts)
    customer_summary = _customer_summary(facts)

    wb = Workbook()
    wb.remove(wb.active)
    _build_dashboard(wb, facts, family_summary, customer_summary, matrix)
    _write_table(wb, "Matriz_RAPV", matrix)
    _write_table(wb, "base_mensual", facts)
    _write_table(wb, "clientes", [_client_row(c) for c in clients])
    _write_table(wb, "productos_equipos", [_product_row(p) for p in products])
    _write_table(wb, "celdas", relationship_cells)
    _write_table(wb, "gav_mensual", gav)
    _write_table(wb, "supuestos", _assumptions())
    _write_table(wb, "diccionario", _dictionary())
    _format_workbook(wb)

    ONEDRIVE_FILE.parent.mkdir(parents=True, exist_ok=True)
    LOCAL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ONEDRIVE_FILE)
    wb.save(LOCAL_FILE)
    print(ONEDRIVE_FILE)
    print(LOCAL_FILE)


def _periods() -> list[str]:
    return [f"{year}-{month:02d}" for year in range(2020, 2026) for month in range(1, 13)]


def _clients() -> list[Client]:
    return [
        Client("CLI-001", "1000001001", "Minera del Sol", "Mineria", "Grupo Minero Sol", "Gran cuenta", "Faena Norte", "Antofagasta"),
        Client("CLI-002", "1000001002", "Minera Cordillera", "Mineria", "Grupo Cordillera", "Gran cuenta", "Faena Cordillera", "Atacama"),
        Client("CLI-003", "1000001003", "Minera Pacifico", "Mineria", "Grupo Pacifico", "Gran cuenta", "Puerto Minero", "Coquimbo"),
        Client("CLI-004", "1000001004", "Constructora Andina", "Construccion", "Grupo Andino", "Cuenta estrategica", "Obras RM", "Metropolitana"),
        Client("CLI-005", "1000001005", "Infraestructura Sur", "Construccion", "Grupo InfraSur", "Cuenta regional", "Proyecto Ruta Sur", "Biobio"),
        Client("CLI-006", "1000001006", "Energia Norte", "Energia", "Grupo Energia Norte", "Cuenta estrategica", "Planta Mejillones", "Antofagasta"),
        Client("CLI-007", "1000001007", "Generacion Austral", "Energia", "Grupo Austral", "Cuenta regional", "Central Los Lagos", "Los Lagos"),
        Client("CLI-008", "1000001008", "Puerto Industrial Chile", "Portuario", "Grupo Puerto Chile", "Cuenta estrategica", "Terminal San Antonio", "Valparaiso"),
        Client("CLI-009", "1000001009", "Servicios Maritimos Valparaiso", "Maritimo", "Grupo Maritimo V", "Cuenta regional", "Puerto Valparaiso", "Valparaiso"),
        Client("CLI-010", "1000001010", "Industria Celulosa BioBio", "Forestal", "Grupo Celulosa Sur", "Cuenta regional", "Planta BioBio", "Biobio"),
        Client("CLI-011", "1000001011", "Agroindustrial Central", "Agroindustria", "Grupo Agro Central", "Cuenta regional", "Planta Central", "Maule"),
        Client("CLI-012", "1000001012", "Data Center Santiago", "Tecnologia", "Grupo DC Chile", "Cuenta estrategica", "Campus Santiago", "Metropolitana"),
    ]


def _products() -> list[Product]:
    return [
        Product("PROD-001", "MAT-001001", "Motores", "Motores estaticos", "Servicios", "Motor", "C175", "Equipo comercializado hoy", "En servicio", 120),
        Product("PROD-002", "MAT-001002", "Motores", "Motores maritimos", "Repuestos", "Motor", "3516", "Equipo no comercializado / legado", "Proximo a renovacion", 144),
        Product("PROD-003", "MAT-002001", "Generadores", "Generadores diesel", "Arriendos", "Generador", "XQ200", "Equipo comercializado hoy", "Vendido actual", 120),
        Product("PROD-004", "MAT-002002", "Generadores", "Generadores respaldo", "Servicios", "Generador", "C32", "Equipo comercializado hoy", "En servicio", 120),
        Product("PROD-005", "MAT-003001", "Soporte electrico", "UPS", "Servicios", "UPS", "UPS-IND", "Equipo no comercializado / legado", "Fin de vida util", 96),
        Product("PROD-006", "MAT-003002", "Soporte electrico", "Tableros electricos", "Repuestos", "Tablero", "TAB-MT", "Equipo comercializado hoy", "En servicio", 96),
    ]


def _cells(clients: list[Client], products: list[Product], rng: Random) -> list[dict[str, object]]:
    cells: list[dict[str, object]] = []
    cell_id = 1
    for client in clients:
        selected = rng.sample(products, k=rng.randint(2, 4))
        for product in selected:
            cells.append(
                {
                    "id_celda": f"CEL-{cell_id:04d}",
                    "id_cliente": client.id_cliente,
                    "id_producto_equipo": product.id_producto_equipo,
                    "codigo_cliente_sap": client.codigo_cliente_sap,
                    "codigo_material_sap": product.codigo_material_sap,
                    "pais": "Chile",
                    "sociedad_sap": "CL01",
                    "industria": client.industria,
                    "grupo_economico": client.grupo_economico,
                    "nombre_cliente": client.nombre_cliente,
                    "faena_operacion": client.faena_operacion,
                    "familia_producto": product.familia_producto,
                    "subfamilia_producto": product.subfamilia_producto,
                    "comercializacion_equipo": product.comercializacion_equipo,
                    "estado_ciclo_vida": product.estado_ciclo_vida,
                    "activo_desde": "2020-01-01",
                    "activo_hasta": "",
                }
            )
            cell_id += 1
    return cells


def _gav_monthly(periods: list[str]) -> list[dict[str, object]]:
    rows = []
    for idx, period in enumerate(periods):
        year = int(period[:4])
        month = int(period[-2:])
        inflation = (1.055 ** (year - 2020))
        season = 1 + 0.03 * sin(month / 12 * 6.283)
        stress = 0.96 if year == 2020 else 1.05 if year == 2022 else 1.02
        gav = 1_680_000_000 * inflation * season * stress
        rows.append(
            {
                "periodo": period,
                "gav_total": round(gav),
                "moneda": "CLP",
                "fuente": "Simulado FI-CO / CeCo",
            }
        )
    return rows


def _facts(
    periods: list[str],
    cells: list[dict[str, object]],
    clients: list[Client],
    products: list[Product],
    gav: list[dict[str, object]],
    rng: Random,
) -> list[dict[str, object]]:
    client_index = {c.id_cliente: c for c in clients}
    product_index = {p.id_producto_equipo: p for p in products}
    gav_by_period = {row["periodo"]: row["gav_total"] for row in gav}
    weights = {cell["id_celda"]: rng.uniform(0.7, 1.35) for cell in cells}
    weight_total = sum(weights.values())
    rows: list[dict[str, object]] = []
    invoice_number = 9_000_000_000

    for p_idx, period in enumerate(periods):
        year = int(period[:4])
        month = int(period[-2:])
        commodity_cycle = {2020: 0.72, 2021: 1.08, 2022: 1.18, 2023: 0.98, 2024: 1.02, 2025: 0.93}[year]
        interest_pressure = {2020: 0.95, 2021: 0.98, 2022: 1.08, 2023: 1.14, 2024: 1.05, 2025: 1.00}[year]
        capex_cycle = commodity_cycle / interest_pressure
        season = 1 + 0.08 * sin(month / 12 * 6.283)

        for cell in cells:
            client = client_index[str(cell["id_cliente"])]
            product = product_index[str(cell["id_producto_equipo"])]
            family_factor = {
                "Motores": 1.18,
                "Generadores": 1.02,
                "Soporte electrico": 0.88,
            }[product.familia_producto]
            industry_factor = {
                "Mineria": 1.24,
                "Construccion": 0.88,
                "Energia": 1.02,
                "Portuario": 0.92,
                "Maritimo": 0.90,
                "Forestal": 0.82,
                "Agroindustria": 0.75,
                "Tecnologia": 1.05,
            }[client.industria]
            lifecycle_factor = {
                "Vendido actual": 1.08,
                "En servicio": 1.0,
                "Proximo a renovacion": 1.18,
                "Fin de vida util": 1.26,
            }[product.estado_ciclo_vida]
            base = 32_000_000 * family_factor * industry_factor * lifecycle_factor * season * rng.uniform(0.75, 1.22)

            ventas_equipos_nuevos = base * capex_cycle * (0.95 if "comercializado hoy" in product.comercializacion_equipo else 0.18)
            ventas_repuestos = base * (0.55 + 0.08 * lifecycle_factor)
            ventas_servicios = base * (0.42 + 0.10 * lifecycle_factor)
            ventas_arriendos = base * (0.34 if product.familia_producto == "Generadores" else 0.10) * (1.25 - min(capex_cycle, 1.2) * 0.22)

            margen_repuestos = ventas_repuestos * 0.34
            margen_servicios = ventas_servicios * 0.29
            margen_arriendos = ventas_arriendos * 0.38
            margen_postventa = margen_repuestos + margen_servicios + margen_arriendos
            gav_asignado = gav_by_period[period] * weights[cell["id_celda"]] / weight_total
            rapv = margen_postventa / gav_asignado if gav_asignado else 0

            dos_base = {
                "Mineria": 58,
                "Construccion": 74,
                "Energia": 62,
                "Portuario": 69,
                "Maritimo": 76,
                "Forestal": 82,
                "Agroindustria": 88,
                "Tecnologia": 52,
            }[client.industria]
            dos = round(max(32, dos_base + (1 - min(rapv, 1.6)) * 12 + rng.uniform(-9, 11)))
            cobranza_alineada = "Si" if dos <= 65 else "No"

            invoice_number += 1
            rows.append(
                {
                    "periodo": period,
                    "anio": year,
                    "mes": month,
                    "id_celda": cell["id_celda"],
                    "numero_factura_simulada": f"FAC-{invoice_number}",
                    "codigo_cliente_sap": client.codigo_cliente_sap,
                    "nombre_cliente": client.nombre_cliente,
                    "industria": client.industria,
                    "grupo_economico": client.grupo_economico,
                    "faena_operacion": client.faena_operacion,
                    "codigo_material_sap": product.codigo_material_sap,
                    "familia_producto": product.familia_producto,
                    "subfamilia_producto": product.subfamilia_producto,
                    "comercializacion_equipo": product.comercializacion_equipo,
                    "estado_ciclo_vida": product.estado_ciclo_vida,
                    "ventas_equipos_nuevos": round(ventas_equipos_nuevos),
                    "ventas_repuestos": round(ventas_repuestos),
                    "ventas_servicios": round(ventas_servicios),
                    "ventas_arriendos": round(ventas_arriendos),
                    "margen_repuestos": round(margen_repuestos),
                    "margen_servicios": round(margen_servicios),
                    "margen_arriendos": round(margen_arriendos),
                    "margen_postventa": round(margen_postventa),
                    "gav_asignado": round(gav_asignado),
                    "ratio_absorcion": round(rapv, 4),
                    "dos": dos,
                    "cobranza_alineada": cobranza_alineada,
                    "moneda": "CLP",
                }
            )
    return rows


def _relationship_cells(
    cells: list[dict[str, object]], facts: list[dict[str, object]]
) -> list[dict[str, object]]:
    """Enriquece la relacion cliente-producto con metricas acumuladas.

    La hoja `celdas` funciona como tabla puente de la tesis. Por eso se agregan
    metricas que permiten conectar la jerarquia de clientes, la familia de
    productos, la facturacion recurrente y el GAV asignado.
    """
    by_cell: dict[str, list[dict[str, object]]] = {}
    for row in facts:
        by_cell.setdefault(str(row["id_celda"]), []).append(row)

    enriched = []
    for cell in cells:
        rows = by_cell.get(str(cell["id_celda"]), [])
        facturacion_servicios_arriendos = sum(
            float(row["ventas_servicios"]) + float(row["ventas_arriendos"]) for row in rows
        )
        gav_total = sum(float(row["gav_asignado"]) for row in rows)
        margen_postventa = sum(float(row["margen_postventa"]) for row in rows)
        periods = len({row["periodo"] for row in rows}) or 1
        enriched.append(
            {
                **cell,
                "facturacion_servicios_arriendos": round(facturacion_servicios_arriendos),
                "gav_mensual_prorrateado_familia": round(gav_total / periods),
                "margen_postventa_total": round(margen_postventa),
                "rapv_periodo_total": round(margen_postventa / gav_total, 4) if gav_total else 0,
            }
        )
    return enriched


def _matrix(
    cells: list[dict[str, object]],
    clients: list[Client],
    products: list[Product],
    facts: list[dict[str, object]],
) -> list[dict[str, object]]:
    key_facts: dict[tuple[str, str], list[dict[str, object]]] = {}
    for row in facts:
        key_facts.setdefault((str(row["codigo_cliente_sap"]), str(row["familia_producto"])), []).append(row)

    rows: list[dict[str, object]] = []
    families = sorted({p.familia_producto for p in products})
    for client in clients:
        record: dict[str, object] = {
            "pais": "Chile",
            "industria": client.industria,
            "grupo_economico": client.grupo_economico,
            "codigo_cliente_sap": client.codigo_cliente_sap,
            "nombre_cliente": client.nombre_cliente,
        }
        for family in families:
            group = key_facts.get((client.codigo_cliente_sap, family), [])
            if group:
                margin = sum(float(r["margen_postventa"]) for r in group)
                gav = sum(float(r["gav_asignado"]) for r in group)
                dos = sum(float(r["dos"]) for r in group) / len(group)
                record[f"RAPV_{family}"] = round(margin / gav, 3) if gav else 0
                record[f"DOS_{family}"] = round(dos, 1)
            else:
                record[f"RAPV_{family}"] = ""
                record[f"DOS_{family}"] = ""
        rows.append(record)
    return rows


def _family_summary(facts: list[dict[str, object]]) -> list[dict[str, object]]:
    return _group_summary(facts, "familia_producto")


def _customer_summary(facts: list[dict[str, object]]) -> list[dict[str, object]]:
    return _group_summary(facts, "nombre_cliente")


def _group_summary(facts: list[dict[str, object]], key: str) -> list[dict[str, object]]:
    grouped: dict[str, list[dict[str, object]]] = {}
    for row in facts:
        grouped.setdefault(str(row[key]), []).append(row)

    out = []
    for name, rows in grouped.items():
        margin = sum(float(r["margen_postventa"]) for r in rows)
        gav = sum(float(r["gav_asignado"]) for r in rows)
        ventas_postventa = sum(float(r["ventas_repuestos"] + r["ventas_servicios"] + r["ventas_arriendos"]) for r in rows)
        dos = sum(float(r["dos"]) for r in rows) / len(rows)
        out.append(
            {
                key: name,
                "ventas_postventa": round(ventas_postventa),
                "margen_postventa": round(margin),
                "gav_asignado": round(gav),
                "ratio_absorcion": round(margin / gav, 3) if gav else 0,
                "dos_promedio": round(dos, 1),
                "cobranza_alineada": "Si" if dos <= 65 else "No",
            }
        )
    return sorted(out, key=lambda r: float(r["margen_postventa"]), reverse=True)


def _build_dashboard(
    wb: Workbook,
    facts: list[dict[str, object]],
    family_summary: list[dict[str, object]],
    customer_summary: list[dict[str, object]],
    matrix: list[dict[str, object]],
) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    margin = sum(float(r["margen_postventa"]) for r in facts)
    gav = sum(float(r["gav_asignado"]) for r in facts)
    rapv = margin / gav if gav else 0
    dos = sum(float(r["dos"]) for r in facts) / len(facts)

    ws["A1"] = "Resiliencia Financiera Chile - MAFI USACH"
    ws["A2"] = "Modelo simulado 2020-2025: jerarquia de clientes x grupos de equipos"
    ws["A4"] = "Indicador"
    ws["B4"] = "Valor"
    kpis = [
        ("Clientes", len({r["codigo_cliente_sap"] for r in facts})),
        ("Celdas cliente-equipo", len({r["id_celda"] for r in facts})),
        ("Registros mensuales", len(facts)),
        ("Margen post-venta", round(margin)),
        ("GAV asignado", round(gav)),
        ("RAPV", round(rapv, 3)),
        ("DOS promedio", round(dos, 1)),
    ]
    for idx, row in enumerate(kpis, 5):
        ws.cell(idx, 1, row[0])
        ws.cell(idx, 2, row[1])

    start = 14
    headers = ["Grupo producto", "Margen post-venta", "GAV asignado", "RAPV", "DOS promedio", "Cobranza alineada"]
    for col, header in enumerate(headers, 1):
        ws.cell(start, col, header)
    for r_idx, row in enumerate(family_summary, start + 1):
        ws.cell(r_idx, 1, row["familia_producto"])
        ws.cell(r_idx, 2, row["margen_postventa"])
        ws.cell(r_idx, 3, row["gav_asignado"])
        ws.cell(r_idx, 4, row["ratio_absorcion"])
        ws.cell(r_idx, 5, row["dos_promedio"])
        ws.cell(r_idx, 6, row["cobranza_alineada"])

    client_start = start + len(family_summary) + 5
    ws.cell(client_start, 1, "Top clientes por margen post-venta")
    client_headers = ["Cliente", "Margen post-venta", "RAPV", "DOS promedio", "Cobranza alineada"]
    for col, header in enumerate(client_headers, 1):
        ws.cell(client_start + 1, col, header)
    for r_idx, row in enumerate(customer_summary[:8], client_start + 2):
        ws.cell(r_idx, 1, row["nombre_cliente"])
        ws.cell(r_idx, 2, row["margen_postventa"])
        ws.cell(r_idx, 3, row["ratio_absorcion"])
        ws.cell(r_idx, 4, row["dos_promedio"])
        ws.cell(r_idx, 5, row["cobranza_alineada"])

    chart = BarChart()
    chart.title = "RAPV por grupo de productos"
    chart.y_axis.title = "RAPV"
    chart.x_axis.title = "Grupo"
    chart.add_data(Reference(ws, min_col=4, min_row=start, max_row=start + len(family_summary)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(family_summary)))
    chart.height = 7
    chart.width = 13
    ws.add_chart(chart, "H4")

    dos_chart = LineChart()
    dos_chart.title = "DOS promedio por grupo de productos"
    dos_chart.y_axis.title = "Dias"
    dos_chart.add_data(Reference(ws, min_col=5, min_row=start, max_row=start + len(family_summary)), titles_from_data=True)
    dos_chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(family_summary)))
    dos_chart.height = 7
    dos_chart.width = 13
    ws.add_chart(dos_chart, "H20")


def _write_table(wb: Workbook, name: str, rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(name)
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _format_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    dashboard_fill = PatternFill("solid", fgColor="1F6FEB")
    white = Font(color="FFFFFF", bold=True)
    header_font = Font(color="17202A", bold=True)
    title_font = Font(color="17202A", bold=True, size=18)
    subtitle_font = Font(color="52616F", size=11)
    thin = Side(style="thin", color="D9E1E8")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
                cell.border = Border(bottom=thin)
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
        if ws.title == "Dashboard":
            ws["A1"].font = title_font
            ws["A2"].font = subtitle_font
            for cell in ws[4]:
                cell.fill = dashboard_fill
                cell.font = white
            for cell in ws[14]:
                cell.fill = dashboard_fill
                cell.font = white
            for col in range(1, 13):
                ws.column_dimensions[get_column_letter(col)].width = 18
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["H"].width = 22
        else:
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = min(28, max(12, len(str(ws.cell(1, col).value)) + 3))
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, (int, float)):
                    continue
                header = ws.cell(1, cell.column).value
                header_text = str(header or "").lower()
                if (
                    "ratio_absorcion" in header_text
                    or "rapv" in header_text
                    or header_text.startswith("rapv_")
                ):
                    cell.number_format = "0.00%"
                elif "dos" in header_text:
                    cell.number_format = "0"
                else:
                    cell.number_format = "#,##0"
        if ws.title == "Dashboard":
            ws["B10"].number_format = "0.00%"
            for row in range(15, 15 + 20):
                if ws.cell(row, 4).value not in (None, ""):
                    ws.cell(row, 4).number_format = "0.00%"
            for row in range(24, 24 + 20):
                if ws.cell(row, 3).value not in (None, ""):
                    ws.cell(row, 3).number_format = "0.00%"


def _client_row(c: Client) -> dict[str, object]:
    return {
        "id_cliente": c.id_cliente,
        "codigo_cliente_sap": c.codigo_cliente_sap,
        "nombre_cliente": c.nombre_cliente,
        "pais": "Chile",
        "sociedad_sap": "CL01",
        "industria": c.industria,
        "grupo_economico": c.grupo_economico,
        "segmento": c.segmento,
        "faena_operacion": c.faena_operacion,
        "region": c.region,
        "estado_cliente": "Activo",
    }


def _product_row(p: Product) -> dict[str, object]:
    return {
        "id_producto_equipo": p.id_producto_equipo,
        "codigo_material_sap": p.codigo_material_sap,
        "familia_producto": p.familia_producto,
        "subfamilia_producto": p.subfamilia_producto,
        "linea_negocio": p.linea_negocio,
        "tipo_equipo": p.tipo_equipo,
        "modelo": p.modelo,
        "comercializacion_equipo": p.comercializacion_equipo,
        "estado_ciclo_vida": p.estado_ciclo_vida,
        "vida_util_esperada_meses": p.vida_util_esperada_meses,
    }


def _assumptions() -> list[dict[str, object]]:
    return [
        {"supuesto": "Periodo", "valor": "2020-01 a 2025-12", "comentario": "72 meses completos para la tesis"},
        {"supuesto": "Pais", "valor": "Chile", "comentario": "Alcance inicial del prototipo"},
        {"supuesto": "RAPV", "valor": "Margen post-venta / GAV asignado", "comentario": "Post-venta = repuestos + servicios + arriendos"},
        {"supuesto": "DOS", "valor": "Days of standing / dias de cobranza", "comentario": "Simulado por industria y resiliencia de celda"},
        {"supuesto": "Cobranza alineada", "valor": "Si cuando DOS <= 65", "comentario": "Umbral inicial para prototipo"},
        {"supuesto": "Datos SAP", "valor": "Simulados", "comentario": "Luego se reemplazan por exportacion VBScript"},
    ]


def _dictionary() -> list[dict[str, object]]:
    return [
        {"campo": "id_celda", "descripcion": "Interseccion cliente x grupo/material de equipo"},
        {"campo": "ratio_absorcion", "descripcion": "Margen post-venta dividido por GAV asignado"},
        {"campo": "margen_postventa", "descripcion": "Margen bruto de repuestos, servicios y arriendos"},
        {"campo": "gav_asignado", "descripcion": "Gasto de administracion y ventas asignado a la celda"},
        {"campo": "facturacion_servicios_arriendos", "descripcion": "Facturacion total de servicios y arriendos acumulada en la celda"},
        {"campo": "gav_mensual_prorrateado_familia", "descripcion": "GAV promedio mensual asignado a la celda dentro de su familia de productos"},
        {"campo": "rapv_periodo_total", "descripcion": "RAPV acumulado de la celda para todo el periodo simulado"},
        {"campo": "dos", "descripcion": "Dias promedio de recuperacion/cobranza"},
        {"campo": "comercializacion_equipo", "descripcion": "Distingue equipos que se venden hoy versus equipos legacy"},
        {"campo": "estado_ciclo_vida", "descripcion": "Estado del ciclo de vida del equipo en cliente"},
    ]


if __name__ == "__main__":
    main()
